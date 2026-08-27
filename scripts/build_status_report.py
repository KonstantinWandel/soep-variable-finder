#!/usr/bin/env python3
"""Single source of truth for per-source progress.

Reads the registry, each source's `raw/` folder and `FETCH_LOG.json`, and the built
`geodb_metadata.json`, optionally re-checks every portal URL, and writes:

  data_sources/CHECKLIST.md          one line per source: state, artifacts, records, next step
  Geospatial_Data_Sources.xlsx       adds a machine-generated "Status_GeoDB" sheet
                                     (all AI-written cells in blue, per the workspace rule);
                                     the untouched original is kept as *_orig.xlsx

Run:
  python scripts/build_status_report.py                 # with a live link check
  python scripts/build_status_report.py --no-link-check
"""
from __future__ import annotations

import argparse
import json
import ssl
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_SOURCES = REPO_ROOT / "data_sources"
REGISTRY = DATA_SOURCES / "registry" / "geo_sources.json"
METADATA = REPO_ROOT / "soep_metadata_output" / "geodb_metadata.json"
WORKBOOK = REPO_ROOT / "Geospatial_Data_Sources.xlsx"
WORKBOOK_ORIGINAL = REPO_ROOT / "Geospatial_Data_Sources_orig.xlsx"
CHECKLIST = DATA_SOURCES / "CHECKLIST.md"
# Clean handoff folder: only what a human needs, named so it is identifiable on sight.
DELIVERABLES = REPO_ROOT / "deliverables_geodb_datenquellen"
PROGRESS_BASE = "geodb_datenquellen_fortschritt"

AI_BLUE = Font(color="FF1F77B4")
AI_BLUE_BOLD = Font(color="FF1F77B4", bold=True)

UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/122.0.0.0 Safari/537.36")

# Which builder source_key(s) each workbook source produces records under.
SOURCE_KEYS: Dict[str, List[str]] = {
    "regionalatlas-deutschland": ["regionalatlas"],
    "datenguide-abgeschaltet": ["regionalstatistik"],
    "strukturdaten-und-indikatoren-ba": ["ba_strukturdaten"],
    "strukturdaten-bundestagswahl-2021": ["btw21_strukturdaten"],
    "migration-integration-in-regionen": ["migration_integration"],
    "hochschulkompass": ["hochschulkompass"],
    "laendermonitor-fruehkindliche-bildungssysteme": ["laendermonitor"],
    "deutschlandatlas-erreichbarkeit-von-apotheken": ["deutschlandatlas"],
    "krankenhausverzeichnis": ["gba_qualitaetsbericht"],
    "bundes-klinik-atlas": ["bundes_klinik_atlas"],
    "open-data-oepnv": ["opendata_oepnv", "transit_formats"],
    "german-companies": ["german_companies"],
    "unfallatlas": ["unfallatlas"],
    "arbeitsmarktreport-ba": ["ba_arbeitsmarktreport"],
    "deutsche-bahn-infrastrukturregister": ["db_isr"],
    "genesis-online-bund": ["genesis_bund"],
    "zensus-2022": ["zensus2022"],
    "breitband-monitor": ["breitband"],
    "arbeitsmarkt-kommunal-ba": ["ba_arbeitsmarkt_kommunal"],
}

# Curated per-source state. `state` drives the checklist marker:
#   done     nothing outstanding
#   partial  indexed, but a better or fuller catalogue is still reachable
#   open     needs a human step before it can be indexed properly
# `next` is the developer detail (English, shown in CHECKLIST.md); `de` is the short phrase
# that goes into the German handoff table.
OPEN_ITEMS: Dict[str, Dict[str, str]] = {
    "regionalatlas-deutschland": {"de": "Nichts offen; Katalog bei Aktualisierung neu ziehen", "state": "done", "next": "Nothing outstanding. Re-fetch services.json when the atlas updates (it carries a timestamp per theme)."},
    "breitband-monitor": {"de": "Eingebunden: Breitbandatlas, Mobilfunk-Monitoring und Rasterdaten (Gitterzellen)", "state": "done", "next": "632 indicators: the Breitbandatlas and Mobilfunk-Monitoring workbooks (use case x technology/bandwidth, Bund to Gemeinde) plus the two GeoPackages read at schema level (3.59 million grid cells x 168 coverage attributes, and 599,515 cells of mobile-operator counts). The GeoPackages themselves are never unpacked into the repo; scripts/extract_gpkg_schema.py writes a small schema JSON instead."},
    "breitbandatlas": {"de": "Nachfolger Gigabitgrundbuch; Indikatorliste fehlt noch", "state": "open", "next": "The bmvi.de link in the workbook is dead; Gigabitgrundbuch is the successor. Needs the indicator/download page saved from a browser."},
    "arbeitsmarktstatistik-ba-karte": {"de": "Glossar eingebunden; Karten-Indikatorliste nicht maschinenlesbar", "state": "partial", "next": "Glossary and portal saved. The map's own indicator list is not machine-readable; BA's API page may expose a catalogue worth crawling."},
    "strukturdaten-und-indikatoren-ba": {"de": "Nichts offen; neueres Heft bei Bedarf", "state": "done", "next": "One booklet defines the series. Refresh with a newer heft when the BA publishes one."},
    "arbeitsmarktreport-ba": {"de": "288 Merkmale aus 17 Blättern eingebunden, monatliche Reihe", "state": "done", "next": "288 indicators flattened from the 17 data sheets (Eckwerte, SGB II/III, Unterbeschäftigung, Alo_Bestand/Bewegungen, Arbeitsstellen, Berufe, Ausbildung, Beschäftigung, Grundsicherung). Labels that recur across sheets carry their sheet in brackets, since 'Bestand an Arbeitslosen: Insgesamt' means something different in Eckwerte and in Eckwerte SGB II."},
    "arbeitsmarkt-kommunal-ba": {"de": "Gemeindescharfe Merkmale eingebunden; weitere Kreis-Hefte optional", "state": "done", "next": "33 indicators flattened from one district archive (one XLSX per municipality, sheet 'Daten'). The indicator set is identical across districts, so more archives add regions, not concepts."},
    "migration-integration-in-regionen": {"de": "Nichts offen", "state": "done", "next": "Nothing outstanding."},
    "krankenhausatlas-deutschland": {"de": "Stand 2016; praktisch ersetzt durch G-BA und Klinik-Atlas", "state": "open", "next": "Portal page only, and the atlas is at 2016. Superseded in practice by the G-BA Qualitätsberichte and the Bundes-Klinik-Atlas, both indexed."},
    "krankenhausverzeichnis": {"de": "Schema mit Unterabschnitten eingebunden (52), Einzelberichte bewusst nicht", "state": "done", "next": "Schema sections indexed from the 2024 archive; the 2008-2024 archives are on disk (about 1.7 GB per year uncompressed, deliberately never extracted). Indexing the per-hospital rows would be a different product."},
    "bundes-klinik-atlas": {"de": "Nichts offen; ersetzt die eingestellte Weisse Liste", "state": "done", "next": "Row renamed in the workbook on 2026-08-25: Weisse Liste is discontinued, the Bundes-Klinik-Atlas open-data export (IQTIG, 1,577 sites with coordinates) replaces it and is indexed."},
    "arztsuche-bundesaerztekammer": {"de": "Nur Suchmaske, kein Export; nur Portaleintrag möglich", "state": "open", "next": "Search UI over the Landesärztekammer registers, no export. Portal-level record only unless a state chamber publishes a list."},
    "deutschlandatlas-erreichbarkeit-von-apotheken": {"de": "Alle Indikatoren eingebunden; Einzel-Links brauchen einen Browser", "state": "partial", "next": "All 86 indicators are indexed from the PDF and XLSX, and this row covers the whole Deutschlandatlas rather than only the pharmacy map. What is NOT possible from here: per-indicator map links. The site answers 400 to every scripted request including its sitemaps, and the PDF carries no map URLs, so the 86 records link at dataset level and are the only ones left with link_verified false. One browser session saving the indicator-to-page mapping would fix it."},
    "hochschulkompass": {"de": "Nichts offen; aktualisierte Liste einfach ersetzbar", "state": "done", "next": "Register attributes indexed. A refreshed hs_liste.txt is a drop-in replacement."},
    "deutsche-bahn-infrastrukturregister": {"de": "Ohne Anmeldung eingebunden: Kartenebenen und WFS-Merkmale des Schienennetzes", "state": "done", "next": "No registration needed, and it does publish a machine-readable catalogue after all. The viewer is a MapStore2 app over a public GeoServer: WMS GetCapabilities lists the map themes (Streckenklasse, Elektrifizierung, ETCS, Gleisanzahl, Betriebsstellen, Bahnsteige, Tunnel, Bruecken, Bahnuebergaenge) and WFS DescribeFeatureType lists the attributes per feature type. Both are indexed, so this row went from one portal card to 436 records. German and English field names are paired where ISR publishes both. Optional next step: DB's StaDa station dataset for the Bahnhofsuche row."},
    "deutsche-bahn-bahnhofsuche": {"de": "Nur Portaleintrag; StaDa-API braucht einen kostenlosen Schlüssel", "state": "open", "next": "Portal page only. The systematic upgrade is DB's StaDa station API (station category, address, facilities, accessibility), which answers HTTP 401 without a key: register free at developers.deutschebahn.com, put the key in ~/kwandel/.config/secrets/ and it can be indexed like the others. Everything else about this row is already covered by the ISR layers."},
    "open-data-oepnv": {"de": "71 Datensätze und die Feldschemata von GTFS und NeTEx eingebunden", "state": "done", "next": "71 named datasets indexed from the public catalogue (Deutschlandweite Sollfahrplandaten GTFS/NeTEX, Deutschlandweite Haltestellendaten, plus Soll-Fahrplandaten/Haltestellen/Liniendaten per Verbund), each with its own deep link, plus the no-login OpenService API products. Downloading a dataset still needs the free account."},
    "spielplatztreff-suchmaschine-fuer-spielplaetze": {"de": "Kein Export; OSM wäre die systematische Alternative", "state": "open", "next": "Crowd-sourced search UI, no export. Portal record only; OSM leisure=playground is the systematic alternative."},
    "spielplatzkarte": {"de": "Kein Export; wie Spielplatztreff", "state": "open", "next": "Same as Spielplatztreff: map UI, no export."},
    "destatis-regionale-mobilitaet-und-infektionsgesc": {"de": "Eingestellt (2020-2022); Indikatoren optional nachtragbar", "state": "partial", "next": "Discontinued experimental statistic (2020-2022). The published mobility indicators could be flattened from the EXSTAT page if the historical series matters."},
    "datenguide-abgeschaltet": {"de": "Merkmalskatalog (2.757) + 866 Regionaltabellen eingebunden", "state": "done", "next": "Two catalogues under this row: the Datenguide GENESIS Merkmalskatalog (2,757 Merkmale) and, since 2026-08-25, the live Regionaldatenbank table catalogue pulled over the API (129 statistics, 866 tables, each with a working table-level deep link). Re-run scripts/fetch_genesis_catalogue.py to refresh. The federal instance (genesis.destatis.de, own token) is not enumerated yet."},
    "inkar": {"de": "Ursprungsquelle des Finders", "state": "done", "next": "Already the finder's original source (660 indicators)."},
    "german-companies": {"de": "Felder eingebunden; kein Bulk-Download möglich (Lookup-API)", "state": "done", "next": "API fields indexed and verified against live responses (samples in raw/). The endpoint is POST /lookup only, a record-linkage service: it resolves a company you already name. There is no bulk or search endpoint (every GET 404s, a city-only filter returns 0 rows), so the register cannot be downloaded through it, and the finder does not need it to. Key now lives in ~/kwandel/.config/secrets/."},
    "open-data-handelsregister": {"de": "Nur Portaleintrag; Feldliste nachtragbar", "state": "partial", "next": "Portal record only. The dump's field list (offeneregister schema) could be flattened without downloading the multi-GB data."},
    "laendermonitor-fruehkindliche-bildungssysteme": {"de": "17 Indikatoren mit offiziellen Methodik-Definitionen eingebunden", "state": "done", "next": "All 17 indicators now carry their official definition, extracted from the four public Methodik PDFs. Those are two-column, so they only read correctly in reading order (pdftotext WITHOUT -layout); with -layout every definition picks up half a sentence from the neighbouring column."},
    "genesis-online-bund": {"de": "3.026 Tabellen über die API eingebunden, Tabellenlinks im Browser geprüft",
        "state": "partial", "next": "3,026 tables enumerated over the REST API (331 statistics) and indexed with table-level links, confirmed by hand in a browser on 2026-08-25. Mostly Bund/Land depth: only 55 titles name Kreise or Gemeinden, which is why this stays partial for a regional finder."},
    "zensus-2022": {"de": "1.440 Tabellen eingebunden, räumliche Ebene je Tabelle aufgelöst (1.407)",
        "state": "partial", "next": "1,440 tables from 12 statistics indexed with table-level links, confirmed by hand in a browser on 2026-08-25. The regional level is encoded in the opaque table code rather than the title; resolved per table through metadata/table on 2026-08-27: 1,407 of 1,440 tables now carry their real level, and a repeated title carries it in the label, which is what tells the four 'Personen: Religion' tables (Bundeslaender, Landeskirche, Bistum, Wahlkreise) apart. Remaining: nearly every table also has a national column, so 'Bund' appears alongside the finer level."},
    "unfallatlas": {"de": "Eingebunden: 25 Merkmale, Unfalljahre 2016-2025, punktgenau", "state": "done", "next": "Attributes of the geocoded accident records indexed (2016-2025, point level with WGS84 and UTM32 coordinates). The yearly CSV archives stay on disk; individual accidents are never indexed."},
    "strukturdaten-bundestagswahl-2021": {"de": "Nichts offen; Strukturdaten 2025 wären eine Erweiterung", "state": "done", "next": "Nothing outstanding. The 2025 Strukturdaten would extend it."},
}


# Sources NOT in the workbook that a German regional-data finder arguably should carry.
# Ordered by what they would add that nothing already indexed provides.
CANDIDATES = [
    ("Zensus 2022", "https://ergebnisse.zensus2022.de/",
     "The one large gap. Buildings, dwellings, households, employment and religion down to the "
     "100 m grid and municipality level, far finer than anything else here. REST API with a token; "
     "the `zensus-genesis-api` skill already documents the header-auth gotcha and a working client."),
    ("Unfallatlas (Destatis)", "https://unfallatlas.statistikportal.de/",
     "Every reported road accident as a geocoded point, 2016 onwards, free download by year and "
     "Land. Point-level accident data exists nowhere else in this list; INKAR only has rates."),
    ("Regionalstatistik table catalogue (live)", "https://www.regionalstatistik.de/genesis/online",
     "We index the GENESIS Merkmale via the Datenguide snapshot. A live crawl of the table "
     "catalogue (statistics -> tables -> regional depth) would add the actual downloadable tables "
     "and their currency. Needs a Regionalstatistik webservice token."),
    ("OpenStreetMap / Overpass POI layers", "https://overpass-turbo.eu/",
     "The systematic replacement for the crowd-sourced portals in the workbook: playgrounds, "
     "pharmacies, GP practices, schools, kindergartens, stops, supermarkets, all as coordinates "
     "with a documented tag schema. Free, no registration, reproducible queries."),
    ("Wegweiser Kommune (Bertelsmann Stiftung)", "https://www.wegweiser-kommune.de/",
     "About 100 indicators for every municipality above 5,000 inhabitants plus demographic "
     "projections to 2040. Complements INKAR on the projection side, which nothing here has."),
    ("BORIS-D / Bodenrichtwerte", "https://www.bodenrichtwerte-boris.de/",
     "Official land values from the Gutachterausschüsse, parcel level. The land-price counterpart "
     "to INKAR's asking rents."),
    ("DWD Climate Data Center", "https://opendata.dwd.de/climate_environment/CDC/",
     "Station and gridded climate series (temperature, precipitation, heat days) at 1 km. The only "
     "environmental/climate axis; free and openly downloadable."),
    ("RWI-GEO-GRID / RWI-GEO-RED (FDZ Ruhr)", "https://fdz.rwi-essen.de/",
     "1 km grid socio-economic data and geocoded real-estate advertisements. Scientific-use files "
     "on application, heavily used in German regional research."),
    ("Election results (Bundeswahlleiter and the Länder)", "https://www.bundeswahlleiter.de/",
     "We index the 2021 structural data but not the results. Constituency and municipality level "
     "results for federal, European and state elections are downloadable as CSV."),
    ("IÖR-Monitor", "https://www.ioer-monitor.de/",
     "Around 90 land-use and landscape-quality indicators at fine spatial resolution, with a WMS/WFS "
     "API. Deeper on land use than the ALKIS shares in INKAR and Regionalatlas."),
]

# How precisely a record's outward link lands on the thing it describes.
LINK_LEVEL_WORD = {
    "indicator": "straight to the indicator",
    "table": "straight to the table",
    "statistic": "to the statistic containing it",
    "dataset": "to the dataset containing it",
    "portal": "to the portal (search from there)",
}

STATE_MARK = {"done": "[x]", "partial": "[~]", "open": "[ ]"}
STATE_WORD = {"done": "done", "partial": "partial", "open": "open"}


def check_url(url: str) -> str:
    if not url:
        return "no url"
    request = urllib.request.Request(url, headers={"User-Agent": UA}, method="GET")
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            return f"HTTP {response.status}"
    except urllib.error.HTTPError as exc:
        return f"HTTP {exc.code}"
    except (urllib.error.URLError, ssl.SSLError, OSError) as exc:
        reason = getattr(exc, "reason", exc)
        return f"unreachable ({str(reason)[:40]})"


def gather(link_check: bool) -> List[Dict[str, Any]]:
    registry = json.loads(REGISTRY.read_text(encoding="utf-8"))["sources"]
    records: List[Dict[str, Any]] = []
    if METADATA.exists():
        records = json.loads(METADATA.read_text(encoding="utf-8"))
    counts: Dict[str, int] = {}
    link_levels: Dict[str, Dict[str, int]] = {}
    unverified: Dict[str, int] = {}
    for record in records:
        key = record["source_key"]
        counts[key] = counts.get(key, 0) + 1
        level = record.get("link_level", "portal")
        link_levels.setdefault(key, {})
        link_levels[key][level] = link_levels[key].get(level, 0) + 1
        if record.get("link_verified") is False:
            unverified[key] = unverified.get(key, 0) + 1

    rows: List[Dict[str, Any]] = []
    for position, source in enumerate(registry, start=1):
        folder = DATA_SOURCES / f"{position:02d}-{source['slug']}"
        raw = folder / "raw"
        files = sorted(
            (p for p in raw.iterdir() if p.name not in {".gitkeep", "FETCH_LOG.json"}),
            key=lambda p: p.name,
        ) if raw.exists() else []
        size = sum(p.stat().st_size for p in files if p.is_file())
        log_path = raw / "FETCH_LOG.json"
        log = json.loads(log_path.read_text(encoding="utf-8")) if log_path.exists() else {"artifacts": {}}
        indexed = sum(counts.get(key, 0) for key in SOURCE_KEYS.get(source["slug"], []))
        levels: Dict[str, int] = {}
        unverified_count = 0
        for key in SOURCE_KEYS.get(source["slug"], []):
            for level, count in link_levels.get(key, {}).items():
                levels[level] = levels.get(level, 0) + count
            unverified_count += unverified.get(key, 0)
        item = OPEN_ITEMS.get(source["slug"], {"state": "open", "next": "not reviewed yet", "de": "noch nicht geprüft"})
        rows.append({
            "position": position,
            "name": source["name"],
            "slug": source["slug"],
            "url": source["url"],
            "folder": folder.name,
            "files": [p.name for p in files],
            "bytes": size,
            "fetched": len(log.get("artifacts", {})),
            "indexed": indexed,
            "portal_record": 1 if source["slug"] != "inkar" else 0,
            "link_levels": levels,
            "unverified_links": unverified_count,
            "state": item["state"],
            "next": item["next"],
            "next_de": item.get("de", item["next"]),
            "note": source["note"],
            "coverage": f"{source['coverage_start_year'] or '?'}-{source['coverage_end_year'] or '?'}"
                        if (source["coverage_start_year"] or source["coverage_end_year"]) else "",
            "access": ", ".join(source["access_modes"]),
            "link": "",
        })

    if link_check:
        with ThreadPoolExecutor(max_workers=8) as pool:
            results = list(pool.map(lambda row: check_url(row["url"]), rows))
        for row, result in zip(rows, results):
            row["link"] = result
    return rows


def human_bytes(size: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024 or unit == "GB":
            return f"{size:.0f} {unit}" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} GB"


def write_checklist(rows: List[Dict[str, Any]], stamp: str) -> None:
    total_indexed = sum(row["indexed"] for row in rows)
    done = sum(1 for row in rows if row["state"] == "done")
    partial = sum(1 for row in rows if row["state"] == "partial")
    open_count = sum(1 for row in rows if row["state"] == "open")

    lines: List[str] = []
    lines.append("# Source checklist")
    lines.append("")
    lines.append(f"Generated by `scripts/build_status_report.py` on {stamp}. Do not hand-edit: "
                 "change `OPEN_ITEMS` in that script and re-run.")
    lines.append("")
    lines.append(f"**{done} done, {partial} partial, {open_count} open** of {len(rows)} sources. "
                 f"{total_indexed} indicator-level records built (plus one portal-level record per source).")
    lines.append("")
    lines.append("`[x]` nothing outstanding | `[~]` indexed, a fuller catalogue is still reachable | "
                 "`[ ]` needs a human step")
    lines.append("")
    totals: Dict[str, int] = {}
    for row in rows:
        for level, count in row["link_levels"].items():
            totals[level] = totals.get(level, 0) + count
    if totals:
        lines.append("**Link precision across all records** (where a hit actually takes the reader): "
                     + ", ".join(f"{count} {LINK_LEVEL_WORD.get(level, level)}"
                                 for level, count in sorted(totals.items(), key=lambda kv: -kv[1])) + ".")
        lines.append("")
    for row in rows:
        lines.append(f"## {STATE_MARK[row['state']]} {row['position']:02d}. {row['name']}")
        lines.append("")
        lines.append(f"- **Folder:** `data_sources/{row['folder']}/`")
        lines.append(f"- **Portal:** {row['url'] or '_none_'}" + (f"  ({row['link']})" if row["link"] else ""))
        lines.append(f"- **Coverage in workbook:** {row['coverage'] or 'not stated'}"
                     + (f" | access: {row['access']}" if row["access"] else "")
                     + (f" | note: {row['note']}" if row["note"] else ""))
        if row["files"]:
            shown = ", ".join(f"`{name}`" for name in row["files"][:6])
            more = f" (+{len(row['files']) - 6} more)" if len(row["files"]) > 6 else ""
            lines.append(f"- **Downloaded:** {len(row['files'])} file(s), {human_bytes(row['bytes'])}: {shown}{more}")
        else:
            lines.append("- **Downloaded:** nothing yet")
        lines.append(f"- **Indexed:** {row['indexed']} indicator-level record(s)"
                     + (" + 1 portal-level record" if row["portal_record"] else ""))
        if row["link_levels"]:
            spelled = ", ".join(f"{count} x {LINK_LEVEL_WORD.get(level, level)}"
                                for level, count in sorted(row["link_levels"].items(), key=lambda kv: -kv[1]))
            if row["unverified_links"]:
                spelled += (f" ({row['unverified_links']} of them not verifiable from here: the target "
                            "portal is a client-rendered app or refuses scripted requests)")
            lines.append(f"- **Link precision:** {spelled}")
        lines.append(f"- **Next step:** {row['next']}")
        lines.append("")
    lines.append("## Candidate sources not in the workbook")
    lines.append("")
    lines.append("Suggested additions, most valuable first. Each would cover something nothing "
                 "currently indexed provides. Nothing here is downloaded yet.")
    lines.append("")
    for name, url, why in CANDIDATES:
        lines.append(f"- **{name}** ({url}): {why}")
    lines.append("")
    CHECKLIST.write_text("\n".join(lines), encoding="utf-8")


def write_workbook(rows: List[Dict[str, Any]], stamp: str) -> None:
    if not WORKBOOK.exists():
        return
    if not WORKBOOK_ORIGINAL.exists():
        WORKBOOK_ORIGINAL.write_bytes(WORKBOOK.read_bytes())

    workbook = openpyxl.load_workbook(WORKBOOK)
    if "Status_GeoDB" in workbook.sheetnames:
        del workbook["Status_GeoDB"]
    sheet = workbook.create_sheet("Status_GeoDB")

    # One language per sheet: the added sheet is English like the rest of the project docs,
    # the original German Tabelle1 is left exactly as it was.
    intro = (f"Integration status for the GeoDB finder (geodb.geolab.soz.uni-bielefeld.de). "
             f"Machine-generated on {stamp} by scripts/build_status_report.py. "
             "Everything on this sheet was added by the assistant (shown in blue); "
             "the original sheet Tabelle1 is untouched.")
    sheet["A1"] = intro
    sheet["A1"].font = AI_BLUE_BOLD
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A1:K1")
    sheet.row_dimensions[1].height = 46

    headers = ["No.", "Data source", "URL", "Link check", "State", "Folder",
               "Downloaded", "Size", "Records indexed", "Link precision", "What is missing / next step"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column, value=title)
        cell.font = AI_BLUE_BOLD
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, row in enumerate(rows, start=3):
        values = [
            row["position"], row["name"], row["url"], row["link"] or "nicht geprüft",
            STATE_WORD[row["state"]], f"data_sources/{row['folder']}/",
            f"{len(row['files'])} file(s)" if row["files"] else "nothing",
            human_bytes(row["bytes"]) if row["bytes"] else "",
            row["indexed"] + row["portal_record"],
            ", ".join(f"{count} x {level}" for level, count in sorted(row["link_levels"].items(), key=lambda kv: -kv[1])) or "portal only",
            row["next"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.font = AI_BLUE
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [5, 34, 46, 22, 12, 34, 16, 10, 12, 26, 70]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"
    workbook.save(WORKBOOK)


STATE_DE = {"done": "fertig", "partial": "teilweise", "open": "offen"}
LINK_LEVEL_DE = {
    "indicator": "Indikator", "table": "Tabelle", "statistic": "Statistik",
    "dataset": "Datensatz", "portal": "Portal",
}


def write_progress_table(rows: List[Dict[str, Any]], stamp: str) -> Optional[Path]:
    """A single presentable table of where every source stands, for handing on. Written as
    CSV and rendered to PDF/PNG through tinytable (one canonical table package, notes inside
    the image), so the deliverable is regenerated by the same command that updates the
    checklist and can never drift from it."""
    import csv as _csv
    import shutil
    import subprocess

    DELIVERABLES.mkdir(parents=True, exist_ok=True)
    csv_path = DELIVERABLES / f"{PROGRESS_BASE}.csv"
    header = ["Nr.", "Datenquelle", "Status", "Dat.", "Anz.", "Verlinkung", "Stand / offener Schritt"]

    def plain(text: str) -> str:
        return (text.replace("\u2265", ">=").replace("\u2264", "<=")
                    .replace("\u2019", "'").replace("\u2018", "'"))
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = _csv.writer(handle)
        writer.writerow(header)
        for row in rows:
            levels = sorted(row["link_levels"].items(), key=lambda kv: -kv[1])
            spelled = ", ".join(f"{LINK_LEVEL_DE.get(level, level)} {count}" for level, count in levels) or "Portal 1"
            if row["unverified_links"]:
                spelled += " (ungeprueft)"
            step = row["next_de"]
            writer.writerow([
                row["position"], plain(row["name"]), STATE_DE[row["state"]],
                len(row["files"]), row["indexed"] + row["portal_record"], spelled,
                plain(row["next_de"]),
            ])

    done = sum(1 for r in rows if r["state"] == "done")
    partial = sum(1 for r in rows if r["state"] == "partial")
    open_count = sum(1 for r in rows if r["state"] == "open")
    total = sum(r["indexed"] + r["portal_record"] for r in rows)
    summary = (f"{len(rows)} Datenquellen: {done} fertig, {partial} teilweise, {open_count} offen; "
               f"{total} indexierte Merkmale im Finder (geodb.geolab.soz.uni-bielefeld.de).")

    rscript = shutil.which("Rscript") or "/home/researcher/miniconda3/envs/rstats/bin/Rscript"
    try:
        subprocess.run([rscript, str(REPO_ROOT / "scripts" / "render_progress_table.R"),
                        str(csv_path), str(DELIVERABLES / PROGRESS_BASE), stamp, summary],
                       check=True, capture_output=True, timeout=600, cwd=str(REPO_ROOT))
    except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        detail = getattr(exc, "stderr", b"")
        print(f"[warn] progress table not rendered: {exc} {detail[-400:] if detail else ''}")
        return csv_path
    return DELIVERABLES / f"{PROGRESS_BASE}.pdf"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--no-link-check", action="store_true")
    parser.add_argument("--date", default=date.today().isoformat(),
                        help="stamp written into the outputs (default: today)")
    args = parser.parse_args()

    rows = gather(link_check=not args.no_link_check)
    write_checklist(rows, args.date)
    write_workbook(rows, args.date)
    progress = write_progress_table(rows, args.date)
    print(json.dumps({
        "sources": len(rows),
        "done": sum(1 for r in rows if r["state"] == "done"),
        "partial": sum(1 for r in rows if r["state"] == "partial"),
        "open": sum(1 for r in rows if r["state"] == "open"),
        "indexed_records": sum(r["indexed"] for r in rows),
        "checklist": str(CHECKLIST),
        "workbook": str(WORKBOOK),
        "progress_table": str(progress) if progress else None,
        "unreachable": [r["name"] for r in rows if r["link"].startswith("unreachable") or r["link"].startswith("HTTP 4") or r["link"].startswith("HTTP 5")],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
